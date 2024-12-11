from agents import company_research_agent
from crewai import Crew, Task
import pandas as pd
from pydantic import BaseModel
from typing import List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import time

countries = [
    "Austria",
    "Belgium",
    "Bulgaria",
    "Croatia",
    "Cyprus",
    "Czech Republic",
    "France",
    "Germany",
    "Greece",
    "Hungary",
    "Italy",
    "Luxembourg",
    "Netherlands",
    "Poland",
    "Romania",
    "Serbia",
    "Slovakia",
    "Slovenia",
    "Sweden",
    "Switzerland",
    "Ukraine",
    "United Arab Emirates",
    "United Kingdom",
    "United States of America"
]

# A dictionary to normalize returned country names to the standardized form in 'countries'
country_normalization_map = {
    "united states": "United States of America",
    "us": "United States of America",
    "u.s.": "United States of America",
    "u.s": "United States of America",
    "usa": "United States of America",
    "uae": "United Arab Emirates",
    "u.a.e.": "United Arab Emirates",
    "united arab emirates": "United Arab Emirates",
    "uk": "United Kingdom",
    "u.k.": "United Kingdom",
    "u.k": "United Kingdom",
    "united kingdom": "United Kingdom"
}

def normalize_country_name(country: str) -> str:
    if not country or not isinstance(country, str):
        return country
    # Strip whitespace and lowercase for matching
    lower_country = country.strip().lower()
    # Replace with the mapped name if available
    normalized = country_normalization_map.get(lower_country, country)
    return normalized

class Company(BaseModel):
    Branches: List[str]
    HQ: str
    Sector_of_Operation: str

company_research_task = Task(
    description="""
        Perform an extensive search to locate all global branches of {company_name} that operates in the sector (if given)
        {sector_of_operation} and its HQ Location.

        A JSON format of all the countries for {company_name}, its HQ location with just country name only
        , and its sector of operation which is basically what industry are they in.

        Ensure that your findings are as recent and accurate as possible.
        The output must be a JSON format of the output. Nothing more and nothing less.
    """,
    agent=company_research_agent,
    expected_output="""It must be a JSON format of the output. Nothing more and nothing less such as the example below.

      "Branches": [
        "United Kingdom",
        "United States of America",
        "Greece"
      ],
      "HQ": "United Arab Emirates",
      "Sector of Operation": "pharma industry"
    """,
    output_pydantic=Company
)

path = r"C:\Users\Asus\Documents\Crowe\crew\clients_list_starter_file.xlsx"

# Load the Excel file into a DataFrame, with headers on the second row
df = pd.read_excel(path, header=1)

# Ensure these columns exist or create them if they do not:
if "Number of Other Countries" not in df.columns:
    df["Number of Other Countries"] = pd.NA
if "Other Countries" not in df.columns:
    df["Other Countries"] = pd.NA

# Columns in Excel (adjust if needed):
hq_column_name = "Country of HQ location"    # Ensure this matches your actual header
sector_column_name = "Sector of operation"   # Ensure this matches your actual header

for index, row in df.iterrows():
    client_name = row["Client name"]
    sector_of_operation = row["Sector of operation"]

    # Prepare inputs for Crew
    inputs = {
        "company_name": client_name,
        "sector_of_operation": sector_of_operation,
    }

    # Run the Crew
    crew = Crew(
        agents=[company_research_agent],
        tasks=[company_research_task],
        memory=False,
        verbose=True,
        cache=False,
    )
    result = crew.kickoff(inputs=inputs)

    # Add a delay after each request to prevent rate limit issues
    time.sleep(15)

    hq = result["HQ"]
    sector = result["Sector_of_Operation"]
    branches = result["Branches"]

    # Normalize the HQ name
    hq = normalize_country_name(hq)
    # Normalize each branch country
    normalized_branches = [normalize_country_name(b) for b in branches]

    # Update HQ column if empty and we have a valid HQ
    if (hq_column_name in df.columns) and (pd.isna(df.at[index, hq_column_name]) or str(df.at[index, hq_column_name]).strip() == "") and hq:
        df.at[index, hq_column_name] = hq

    # Update Sector column if empty and we have a valid sector
    if (sector_column_name in df.columns) and (pd.isna(df.at[index, sector_column_name]) or str(df.at[index, sector_column_name]).strip() == "") and sector:
        df.at[index, sector_column_name] = sector

    # Identify which branches are in the known countries list
    known_branch_countries = set(countries)
    known_branches = [b for b in normalized_branches if b in known_branch_countries]
    unknown_branches = [b for b in normalized_branches if b not in known_branch_countries]

    # For known branches, mark "1" in the respective country columns if found
    for country in countries:
        if (country in df.columns) and (country in known_branches) and (pd.isna(df.at[index, country]) or str(df.at[index, country]).strip() == ""):
            df.at[index, country] = "1"

    # For unknown branches, place the count in "Number of Other Countries" and the names in "Other Countries"
    if unknown_branches:
        df.at[index, "Number of Other Countries"] = len(unknown_branches)
        df.at[index, "Other Countries"] = ", ".join(unknown_branches)
    else:
        df.at[index, "Number of Other Countries"] = pd.NA
        df.at[index, "Other Countries"] = pd.NA

# Replace <NA> values with empty strings for columns that may contain them
df["Number of Other Countries"] = df["Number of Other Countries"].fillna("")
df["Other Countries"] = df["Other Countries"].fillna("")

# Now use openpyxl to preserve formatting
wb = load_workbook(path)
sheet = wb.active

# If headers start at row 2, data starts at row 3 in Excel
for index, row in df.iterrows():
    excel_row = index + 3

    # Write HQ if set
    if hq_column_name in df.columns:
        hq_val = row[hq_column_name]
        # hq_val is now guaranteed to not be <NA>
        sheet["D" + str(excel_row)] = hq_val

    # Write Sector if set
    if sector_column_name in df.columns:
        sector_val = row[sector_column_name]
        sheet["H" + str(excel_row)] = sector_val

    # Write known countries (K to AH)
    start_col = 11  # K=11
    for idx, country in enumerate(countries):
        if country in df.columns:
            country_val = row[country]
            sheet[get_column_letter(start_col + idx) + str(excel_row)] = country_val

    # Write the number of other countries to AI (35th column, after AH which is 34)
    sheet["AI" + str(excel_row)] = row["Number of Other Countries"]

    # Write the other countries to AJ (36th column)
    sheet["AJ" + str(excel_row)] = row["Other Countries"]

wb.save(r'C:\Users\Asus\Documents\Crowe\crew\clients_test_updated.xlsx')

print("Processing complete. The updated file with preserved formatting has been saved.")
